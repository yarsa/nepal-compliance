# openpyxl 3.1.5 - MIT License, (c) 2024, Eric Gazoni, Charlie Clark, See License at https://pypi.org/project/openpyxl/

import frappe
from frappe import _
import json
import openpyxl
from frappe.utils import get_site_path
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from nepal_compliance.nepali_date_utils.utils import bs_date

def convert_to_nepali_fy_format(year_start_date, year_end_date):
    try:
        start_year = year_start_date.year
        end_year = year_end_date.year

        nep_start = start_year + 57
        nep_end = end_year + 57

        return f"{nep_start}/{str(nep_end)[-2:]}"
    except Exception as e:
        return f"{year_start_date.year}-{year_end_date.year}"

@frappe.whitelist()
def generate_ird_sales_register_excel():
    from nepal_compliance.nepal_compliance.report.sales_register_ird.sales_register_ird import get_data

    filters = frappe._dict(json.loads(frappe.form_dict.get("filters") or "{}"))
    rows = get_data(filters)

    if not rows:
        frappe.throw(_("No data found for the selected filters."))

    company = filters.get("company")
    company_info = frappe.get_doc("Company", company) if company else None
    company_name = company_info.company_name if company_info else "Company Name"
    # address = frappe.db.get_value("Address", {"is_your_company_address": 1}, "address_line1") or ""
    pan = company_info.tax_id or "N/A"
    if not rows or len(rows) == 0:
        frappe.throw(_("No data found for the selected filters."))

    invoice_name = rows[0].get("invoice")
    if not invoice_name:
        frappe.throw(_("Missing invoice number in report rows."))
    posting_date = frappe.db.get_value("Sales Invoice", invoice_name, "posting_date")
    if not posting_date:
        frappe.throw(_("Sales Invoice {0} not found").format(invoice_name))
    fy = frappe.db.get_value(
        "Fiscal Year",
        {
            "year_start_date": ["<=", posting_date],
            "year_end_date":   [">=", posting_date]
        },
        ["name", "year_start_date", "year_end_date"],
        as_dict=True,
    )
    if not fy:
        frappe.throw(_("No Fiscal Year found for posting date {0}.").format(posting_date))

    year_start_date = fy.year_start_date
    year_end_date   = fy.year_end_date

    fiscal_year_nepali = convert_to_nepali_fy_format(year_start_date, year_end_date)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Register"

    bold_center = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def format_cell(cell):
        cell.alignment = center
        cell.font = bold_center
        cell.border = border

    ws.merge_cells("A1:L1")
    ws["A1"] = "बिक्री खाता"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A2:L2")
    ws["A2"] = "(नियम २३ को उपनियम (१) को खण्ड  (छ) संग सम्बन्धित )"
    ws["A2"].font = Font(bold=False)
    ws.merge_cells("A3:L3")
    ws["A3"] = ""
    ws.merge_cells("A4:L4")
    ws["A4"] = f"करदाता दर्ता नं (PAN): {pan}        करदाताको नाम: {company_name}         आर्थिक वर्ष: {fiscal_year_nepali}"

    for r in range(1, 5):
        top_cell = ws.cell(row=r, column=1)
        top_cell.font = bold_center
        top_cell.alignment = center

    top_headers = [
        ("बीजक", 4),
        ("जम्मा बिक्री / निकासी (रु)", 1),
        ("स्थानीय कर छुटको बिक्री  मूल्य (रु)", 1),
        ("करयोग्य बिक्री", 2),
        ("निकासी", 4)
    ]

    sub_headers = [
        "मिति", "बीजक नम्बर", "खरिदकर्ताको नाम", "खरिदकर्ताको स्थायी लेखा नम्बर",
        "मूल्य (रु)", "कर (रु)",
        "निकासी गरेको वस्तु वा सेवाको मूल्य (रु)", "निकासी गरेको देश",
        "निकासी प्रज्ञापनपत्र नम्बर", "निकासी प्रज्ञापनपत्र मिति"
    ]

    col = 1
    sub_idx = 0

    for title, span in top_headers:
        if span == 1:
            ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
            cell = ws.cell(row=5, column=col, value=title)
            format_cell(cell)
            col += 1
        else:
            ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + span - 1)
            cell = ws.cell(row=5, column=col, value=title)
            format_cell(cell)
            for i in range(span):
                sub_cell = ws.cell(row=6, column=col + i, value=sub_headers[sub_idx])
                format_cell(sub_cell)
                sub_idx += 1
            col += span

    data_start_row = 7
    for row_idx, inv in enumerate(rows, start=data_start_row):
        posting_date = inv.get("posting_date")
        posting_bs = bs_date(posting_date) if posting_date else ""
        row_data = [
            posting_bs,
            inv.get("invoice"),
            inv.get("customer_name"),
            inv.get("pan"),
            inv.get("total"),
            inv.get("tax_exempt"),
            inv.get("taxable_amount"),
            inv.get("tax_amount"),
            inv.get("Value of Exported Goods or Services"),
            inv.get("export_country"),
            inv.get("Export Declaration Number"),
            inv.get("Export Declaration Date")
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center
            cell.border = border
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'

    total_row = len(rows) + data_start_row
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    total_label_cell = ws.cell(row=total_row, column=1, value="Total")
    format_cell(total_label_cell)

    for col in range(5, 10):
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{total_row - 1})"
        total_cell = ws.cell(row=total_row, column=col, value=formula)
        total_cell.border = border
        total_cell.alignment = center
        total_cell.number_format = '#,##0.00'

    for col in range(10, 13): 
        cell = ws.cell(row=total_row, column=col, value="")
        cell.border = border
        cell.alignment = center

    for col_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 4

    file_name = "IRD_Sales_Register.xlsx"
    path = get_site_path("public", "files", file_name)
    wb.save(path)
    return f"/files/{file_name}"
