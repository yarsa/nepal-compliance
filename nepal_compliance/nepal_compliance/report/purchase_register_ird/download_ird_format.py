# openpyxl 3.1.5 - MIT License, (c) 2024, Eric Gazoni, Charlie Clark, See License at https://pypi.org/project/openpyxl/

import frappe
from frappe import _
from frappe.utils import flt
import json
import openpyxl
from frappe.utils import get_site_path
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

def convert_to_nepali_fy_format(fy_name):
    """
    Convert a fiscal year string from Gregorian format to Nepali fiscal year format.
    
    If the input is in "YYYY-YYYY" format, returns it as "YYYY/YY" with each year incremented by 57. If the input already contains a "/" or conversion fails, returns the original string.
    
    Parameters:
        fy_name (str): Fiscal year in "YYYY-YYYY" or similar format.
    
    Returns:
        str: Fiscal year in Nepali format "YYYY/YY" or the original string if conversion is not possible.
    """
    if "/" in fy_name and len(fy_name.split("/")[0]) == 4:
        return fy_name
    try:
        start, end = [int(x) for x in fy_name.split("-")]
        nep_start = start + 57
        nep_end = end + 57
        return f"{nep_start}/{str(nep_end)[-2:]}"
    except Exception:
        return fy_name


@frappe.whitelist()
def generate_ird_purchase_register_excel():
    """
    Generates and saves an IRD-compliant purchase register Excel report based on filtered purchase data.
    
    The report is formatted according to Nepalese tax authority requirements, including Nepali fiscal year formatting, merged headers in Nepali, and columnar totals (excluding the Permanent Account Number column). The resulting Excel file is saved to the public files directory and the relative download path is returned.
    
    Returns:
        str: The relative URL path to the generated Excel file.
    """
    from nepal_compliance.nepal_compliance.report.purchase_register_ird.purchase_register_ird import get_data

    filters = frappe._dict(json.loads(frappe.form_dict.get("filters") or "{}"))
    rows = get_data(filters)

    if not rows:
        frappe.throw(_("No data found for the selected filters."))

    company = filters.get("company")
    company_info = frappe.get_doc("Company", company) if company else None
    company_name = company_info.company_name if company_info else "Company Name"

    address = frappe.db.get_value(
        "Address",
        {"is_your_company_address": 1},
        "address_line1"
    ) or ""

    pan = company_info.tax_id or "N/A"
    invoice_name = frappe.db.get_value("Purchase Invoice", {"bill_no": rows[0].get("invoice")}, "name") or rows[0].get("invoice")
    invoice_doc = frappe.get_doc("Purchase Invoice", invoice_name)
    
    posting_date = invoice_doc.posting_date

    fiscal_year = frappe.db.get_value(
        "Fiscal Year",
        {
            "year_start_date": ["<=", posting_date],
            "year_end_date": [">=", posting_date]
        },
        "name"
    ) or "Fiscal Year"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Register"

    bold_center = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def format_cell(cell):
        """
        Apply bold font, centered alignment, and a thin border to the given Excel cell.
        """
        cell.alignment = center
        cell.font = bold_center
        cell.border = border

    ws.merge_cells("A1:M1")
    ws["A1"] = f"खरिद खाता"
    ws["A1"].font = Font(bold=True, size=16)

    ws.merge_cells("A2:M2")
    ws["A2"] = f"(नियम २३ को उपनियम (१) को खण्ड  (छ) संग सम्बन्धित )"
    ws["A2"].font = Font(bold=False)

    ws.merge_cells("A3:M3")
    ws["A3"] = ""

    ws.merge_cells("A4:M4")
    ws["A4"] = f"करदाता दर्ता नं (PAN): {pan}        करदाताको नाम: {company_name}         आर्थिक वर्ष: {convert_to_nepali_fy_format(fiscal_year)}"
    ws["A4"].font = bold_center
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r in range(1, 5):
        ws[f"A{r}"].font = bold_center
        ws[f"A{r}"].alignment = center

    # Top headers
    top_headers = [
        ("बीजक", 5),
        ("जम्मा खरिद मूल्य (रु)", 1),
        ("कर छुट हुने वस्तु वा सेवाको खरिद / पैठारी मूल्य (रु)", 1),
        ("करयोग्य खरिद (पूंजीगत बाहेक)", 2),
        ("करयोग्य पैठारी (पूंजीगत बाहेक)", 2),
        ("पूंजीगत करयोग्य खरिद / पैठारी", 2)
    ]

    col = 1
    for title, span in top_headers:
        if span == 1:
            ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
        else:
            ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + span - 1)
        cell = ws.cell(row=5, column=col, value=title)
        format_cell(cell)
        col += span

    sub_headers = {
        1: "मिति",
        2: "बीजक नं.",
        3: "प्रज्ञापनपत्र नं.",
        4: "आपूर्तिकर्ताको नाम",
        5: "आपूर्तिकर्ताको स्थायी लेखा नम्बर",
        6: "मूल्य (रु)",
        7: "मूल्य (रु)",
        8: "मूल्य (रु)",
        9: "कर (रु)",
        10: "मूल्य (रु)",
        11: "कर (रु)",
        12: "मूल्य (रु)",
        13: "कर (रु)"
    }

    for col_num in range(1, 14):
        value = sub_headers.get(col_num)
        if not value:
            continue

        coord = ws.cell(row=6, column=col_num).coordinate
        is_top_left = True
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range and coord != merged_range.start_cell.coordinate:
                is_top_left = False
                break

        if is_top_left:
            cell = ws.cell(row=6, column=col_num, value=value)
            format_cell(cell)

    # Fill data
    for row_idx, inv in enumerate(rows, start=7):
        row_data = [
            inv.get("nepali_date"),
            inv.get("invoice"),
            inv.get("customs_declaration_number"),
            inv.get("supplier_name"),
            inv.get("pan"),
            inv.get("total"),
            inv.get("tax_exempt"),
            inv.get("taxable_amount"),
            inv.get("tax_amount"),
            inv.get("taxable_import_non_capital_amount"),
            inv.get("taxable_import_non_capital_tax"),
            inv.get("capital_taxable_amount"),
            inv.get("capital_taxable_tax")
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center
            cell.border = border
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'

    # Exclude column 5 (Permanent Account Number) from total calculation
    total_row = len(rows) + 7
    ws.cell(row=total_row, column=1, value="Total")
    format_cell(ws.cell(row=total_row, column=1))
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)

    # Calculate totals for other columns, excluding column 5
    for col in range(1, 14):
        if col == 5:
            continue  # Skip column 5
        col_letter = get_column_letter(col)
        col_total = f"=SUM({col_letter}7:{col_letter}{total_row - 1})"
        # Skip merged cells: only set total for non-merged cells
        if not any([coord in merged_range for merged_range in ws.merged_cells.ranges for coord in [ws.cell(row=total_row, column=col).coordinate]]):
            cell = ws.cell(row=total_row, column=col, value=col_total)
            cell.border = border
            cell.alignment = center
            cell.number_format = '#,##0.00'

    # Ensure column 5 (Permanent Account Number) stays black and not part of total calculation
    for row_idx in range(7, total_row):
        ws.cell(row=row_idx, column=5).font = Font(color="000000")  # Black font for column 5

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    # Save the Excel file
    file_name = "IRD_Purchase_Register.xlsx"
    path = get_site_path("public", "files", file_name)
    wb.save(path)

    return f"/files/{file_name}"
