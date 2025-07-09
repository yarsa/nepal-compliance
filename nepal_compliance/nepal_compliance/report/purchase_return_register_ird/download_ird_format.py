# openpyxl 3.1.5 - MIT License, (c) 2024, Eric Gazoni, Charlie Clark, See License at https://pypi.org/project/openpyxl/

import frappe
from frappe import _
from frappe.utils import flt
import json
import openpyxl
from frappe.utils import get_site_path
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

def convert_to_nepali_fy_format(fy_name):
    if "/" in fy_name and len(fy_name.split("/")[0]) == 4:
        return fy_name
    try:
        start, end = [int(x) for x in fy_name.split("-")]
        nep_start = start + 57
        nep_end = end + 57
        return f"{nep_start}/{str(nep_end)[-2:]}"
    except Exception:
        return fy_name

@frappe.whitelist()
def generate_ird_purchase_register_excel():
    from nepal_compliance.nepal_compliance.report.purchase_return_register_ird.purchase_return_register_ird import get_data

    filters = frappe._dict(json.loads(frappe.form_dict.get("filters") or "{}"))
    rows = get_data(filters)

    if not rows:
        frappe.throw(_("No data found for the selected filters."))

    company = filters.get("company")
    company_info = frappe.get_doc("Company", company) if company else None
    company_name = company_info.company_name if company_info else "Company Name"

    # address = frappe.db.get_value("Address", {"is_your_company_address": 1}, "address_line1") or ""
    pan = company_info.tax_id or "N/A"

    invoice_name = frappe.db.get_value("Purchase Invoice", {"bill_no": rows[0].get("invoice")}, "name") or rows[0].get("invoice")
    invoice_doc = frappe.get_doc("Purchase Invoice", invoice_name)
    posting_date = invoice_doc.posting_date

    fiscal_year = frappe.db.get_value("Fiscal Year", {
        "year_start_date": ["<=", posting_date],
        "year_end_date": [">=", posting_date]
    }, "name") or "Fiscal Year"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Return Register"

    bold_center = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    def format_cell(cell):
        cell.alignment = center
        cell.font = bold_center
        cell.border = border

    ws.merge_cells("A1:P1")
    ws["A1"] = "खरिद फिर्ता खाता"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A2:P2")
    ws["A2"] = "(नियम २३ को उपनियम (१) को खण्ड  (छ) संग सम्बन्धित )"
    ws["A2"].font = Font(bold=False)
    ws.merge_cells("A3:P3")
    ws["A3"] = ""
    ws.merge_cells("A4:P4")
    ws["A4"] = f"करदाता दर्ता नं (PAN): {pan}        करदाताको नाम: {company_name}         आर्थिक वर्ष: {convert_to_nepali_fy_format(fiscal_year)}"

    for r in range(1, 5):
        ws[f"A{r}"].alignment = center
        ws[f"A{r}"].font = bold_center

    top_headers = [
        ("बीजक / प्रज्ञापनपत्र नम्बर", 8),
        ("जम्मा फिर्ता मूल्य (रु)", 1),
        ("कर छुट हुने वस्तु वा सेवाको फिर्ता मूल्य (रु)", 1),
        ("करयोग्य फिर्ता (पूंजीगत बाहेक)", 2),
        ("करयोग्य पैठारी फिर्ता (पूंजीगत बाहेक)", 2),
        ("पूंजीगत करयोग्य फिर्ता", 2)
    ]

    col = 1
    skip_subheader = []

    for title, span in top_headers:
        start_col = col
        end_col = col + span - 1
        if span == 1:
            ws.merge_cells(start_row=5, start_column=start_col, end_row=6, end_column=start_col)
            skip_subheader.append(start_col)
        else:
            ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
        cell = ws.cell(row=5, column=start_col, value=title)
        format_cell(cell)
        col += span

    sub_headers = [
        "मिति", "बीजक नं.", "प्रज्ञापनपत्र नं.", "आपूर्तिकर्ताको नाम", "आपूर्तिकर्ताको स्थायी लेखा नम्बर",
        "खरिद/पैठारी फिर्ता गरिएका वस्तु वा सेवाको विवरण", "खरिद/पैठारी फिर्ता गरिएका वस्तु वा सेवाको परिमाण", "वस्तु वा सेवाको एकाई",
        "जम्मा फिर्ता मूल्य (रु)", "कर छुट फिर्ता मूल्य (रु)",
        "मूल्य (रु)", "कर (रु)", "मूल्य (रु)", "कर (रु)", "मूल्य (रु)", "कर (रु)"
    ]

    for col_num, header in enumerate(sub_headers, 1):
        if col_num in skip_subheader:
            continue
        cell = ws.cell(row=6, column=col_num, value=header)
        format_cell(cell)

    data_start_row = 7
    for i, inv in enumerate(rows):
        row_idx = data_start_row + i
        row_data = [
            inv.get("nepali_date"),
            inv.get("invoice"),
            inv.get("customs_declaration_number"),
            inv.get("supplier_name"),
            inv.get("pan"),
            inv.get("reason"),
            inv.get("qty"),
            inv.get("uom"),
            inv.get("total"),
            inv.get("tax_exempt"),
            inv.get("taxable_amount"),
            inv.get("tax_amount"),
            inv.get("taxable_import_non_capital_amount"),
            inv.get("taxable_import_non_capital_tax"),
            inv.get("capital_taxable_amount"),
            inv.get("capital_taxable_tax")
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center
            cell.border = border
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'

    total_row = data_start_row + len(rows)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    total_label_cell = ws.cell(row=total_row, column=1, value="Total")
    format_cell(total_label_cell)

    for col in range(7, 17):
        if col == 8:
            continue
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{total_row - 1})"
        cell = ws.cell(row=total_row, column=col, value=formula)
        cell.alignment = center
        cell.border = border
        cell.number_format = '#,##0.00'

    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 4

    file_name = "IRD_Purchase_Return_Register.xlsx"
    path = get_site_path("public", "files", file_name)
    wb.save(path)

    return f"/files/{file_name}"
