# openpyxl 3.1.5 - MIT License, (c) 2024, Eric Gazoni, Charlie Clark, See License at https://pypi.org/project/openpyxl/

import frappe
from frappe import _
from frappe.utils import get_site_path
import json
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

def convert_to_nepali_fy_format(fy_name):
    if "/" in fy_name and len(fy_name.split("/")[0]) == 4:
        return fy_name
    try:
        start, end = [int(x) for x in fy_name.split("-")]
        nep_start = start + 57
        nep_end = end + 57
        return f"{nep_start}/{str(nep_end)[-2:]}"
    except Exception:
        return fy_name

@frappe.whitelist()
def generate_ird_sales_register_excel():
    from nepal_compliance.nepal_compliance.report.sales_return_register_ird.sales_return_register_ird import get_data

    filters = frappe._dict(json.loads(frappe.form_dict.get("filters") or "{}"))
    rows = get_data(filters)

    if not rows:
        frappe.throw(_("No data found for the selected filters."))

    company = filters.get("company")
    company_info = frappe.get_doc("Company", company) if company else None
    company_name = company_info.company_name if company_info else "Company Name"
    pan = company_info.tax_id or "N/A"

    posting_date = frappe.utils.nowdate()
    fiscal_year = frappe.db.get_value("Fiscal Year", {
        "year_start_date": ["<=", posting_date],
        "year_end_date": [">=", posting_date]
    }, "name") or "Fiscal Year"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Return Register"

    bold_center = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    def format_cell(cell):
        cell.alignment = center
        cell.font = bold_center
        cell.border = border

    ws.merge_cells("A1:K1")
    ws["A1"] = "बिक्री फिर्ता खाता"
    ws["A1"].font = Font(bold=True, size=16)

    ws.merge_cells("A2:K2")
    ws["A2"] = "(नियम २३ को उपनियम (१) को खण्ड  (छ) संग सम्बन्धित )"

    ws.merge_cells("A3:K3")
    ws["A3"] = ""

    ws.merge_cells("A4:K4")
    ws["A4"] = f"करदाता दर्ता नं (PAN): {pan}        करदाताको नाम: {company_name}         आर्थिक वर्ष: {convert_to_nepali_fy_format(fiscal_year)}"
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["A4"].font = bold_center

    for r in range(1, 5):
        ws[f"A{r}"].alignment = center
        ws[f"A{r}"].font = bold_center

    top_headers = [
        ("बीजक", 7),
        ("जम्मा फिर्ता (रु)", 1),
        ("स्थानीय कर छुटको फिर्ता मूल्य (रु)", 1),
        ("करयोग्य फिर्ता", 2),
    ]

    sub_headers = [
        "मिति", "बीजक नं.", "खरिदकर्ताको नाम", "खरिदकर्ताको स्थायी लेखा नम्बर",
        "वस्तु वा सेवाको नाम", "वस्तु वा सेवाको परिमाण", "वस्तु वा सेवाको एकाइ",
        "मूल्य (रु)", "कर छुट (रु)", "मूल्य (रु)", "कर (रु)"
    ]

    col = 1
    sub_idx = 0
    for title, span in top_headers:
        if span == 1:
            ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
            cell = ws.cell(row=5, column=col, value=title)
            format_cell(cell)
            col += 1
        else:
            ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + span - 1)
            cell = ws.cell(row=5, column=col, value=title)
            format_cell(cell)
            for i in range(span):
                sub_cell = ws.cell(row=6, column=col + i, value=sub_headers[sub_idx])
                format_cell(sub_cell)
                sub_idx += 1
            col += span

    data_start_row = 7
    for row_idx, inv in enumerate(rows, start=data_start_row):
        is_grand_total_row = inv.get("customer_name") == "कुल जम्मा"

        if is_grand_total_row:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            total_label_cell = ws.cell(row=row_idx, column=1, value="कुल जम्मा")
            format_cell(total_label_cell)

            values = [
                inv.get("name"), inv.get("qty"), inv.get("uom"),
                inv.get("total"), inv.get("tax_exempt"),
                inv.get("taxable_amount"), inv.get("tax_amount")
            ]
            start_col = 5
        else:
            values = [
                inv.get("nepali_date"), inv.get("invoice"), inv.get("customer_name"), inv.get("pan"),
                inv.get("name"), inv.get("qty"), inv.get("uom"),
                inv.get("total"), inv.get("tax_exempt"),
                inv.get("taxable_amount"), inv.get("tax_amount")
            ]
            start_col = 1

        for offset, val in enumerate(values):
            col_idx = start_col + offset
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center
            cell.border = border
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            if is_grand_total_row:
                cell.font = Font(bold=True)

    last_data_row = data_start_row + len(rows)
    for col in range(1, 12):
        max_len = 0
        for row in range(1, last_data_row):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 4

    filepath = get_site_path("public", "files", "IRD_Sales_Return_Register.xlsx")
    wb.save(filepath)
    return "/files/IRD_Sales_Return_Register.xlsx"
